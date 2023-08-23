import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# File paths
CSV_FILE_PATH = "C:/Users/Cameron.Francis/Documents/TAE working/UOM may 2023.csv"
TAE_TEMPLATE_PATH = "C:/Users/Cameron.Francis/Documents/TAE working/TAE template revised targets.xlsx"

# Helper function to clean up data
def clean_data(df):
    # Convert the CSV file to an ordered DataFrame
    df = pd.read_csv(CSV_FILE_PATH, delimiter=";")
    df["Number of values"] = pd.to_numeric(df["Number of values"], errors="coerce")

    # Keep only the data with a higher "Number of values" for tests with the same "Instrument" and "Test" but different "QC lot no."
    df = df.loc[df.groupby(["Instrument", "Test", "QC"])["Number of values"].idxmax()]

    # Sort the DataFrame based on the original index
    df = df.sort_index()
    # Define Virology and POCT and Liason test keywords
    virology_keywords = [
        "CMV", "cytomegalovirus", "Epstein Barr Virus", "Hep", "hepatitis", "Herpes", "HIV",
        "Rubella", "Syphyllis", "toxoplasma", "varicella", "SYPHILIS", "Measles", "Anti-SARS-CoV-2 S Qaun"
    ]
    poct_keywords = [
        "POC", "CREATP", "INDEX"
    ]
    #liason_keywords = [
    #     ALDO", "Human growth hormone", "growth factor", "Renin", "INDEX"
    #]
    # Remove all Virology and POCT and liaison tests
    df = df[~df["Test"].str.contains('|'.join(virology_keywords), case=False)]
    df = df[~df["Test"].str.contains('|'.join(poct_keywords), case=False)]
    #df = df[~df["Test"].str.contains('|'.join(liason_keywords), case=False)]

    return df

# Load the CSV file and clean up the data
df = pd.read_csv(CSV_FILE_PATH)
df = clean_data(df)

# Load the TAE template
book = load_workbook(TAE_TEMPLATE_PATH)

# Add the DataFrame to the "Raw Data" sheet
raw_data_sheet = book["Raw Data"]

# Write the headers
for col_num, header in enumerate(df.columns, 1):
    raw_data_sheet.cell(row=1, column=col_num, value=header)

# Write the data
for index, row in enumerate(df.itertuples(index=False), start=2):
    for col_num, value in enumerate(row, 1):
        col_letter = get_column_letter(col_num)
        raw_data_sheet[f"{col_letter}{index}"] = value

# Make sure at least one sheet is visible
visible_sheets = [sheet for sheet in book.worksheets if not sheet.sheet_state == 'hidden']
if not visible_sheets:
    book.active.sheet_state = 'visible'

# Save the changes to the TAE template
book.save(TAE_TEMPLATE_PATH)