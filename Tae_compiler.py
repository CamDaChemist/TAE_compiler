import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import filedialog
from openpyxl.styles import PatternFill

def select_file(title="Select a file", filetypes=[("All files", "*.*")]):
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(title=title, filetypes=filetypes)
    return file_path

# File paths
CSV_FILE_PATH = select_file(title="Select the CSV file", filetypes=[("CSV files", "*.csv")])
TAE_TEMPLATE_PATH = select_file(title="Select the TAE template file", filetypes=[("Excel files", "*.xlsx")])


def clean_data(df):
    # Convert the CSV file to an ordered DataFrame
    df = pd.read_csv(CSV_FILE_PATH, delimiter=";")
    df["Number of values"] = pd.to_numeric(df["Number of values"], errors="coerce")

    # Keep only the data with a higher "Number of values" for tests with the same "Instrument" and "Test" but different "QC lot no."
    df = df.loc[df.groupby(["Instrument", "Test", "QC"])["Number of values"].idxmax()]

    # Sort the DataFrame based on the original index
    df = df.sort_index()
    # Define Virology and POCT and Liason test keywords
    virology_keywords = [
        "CMV", "cytomegalovirus", "Epstein Barr Virus", "Hep", "hepatitis", "Herpes", "HIV",
        "Rubella", "Syphyllis", "toxoplasma", "varicella", "SYPHILIS", "Measles", "Anti-SARS-CoV-2 S Qaun"
    ]
    poct_keywords = [
        "POC", "CREATP", "INDEX"
    ]
    liason_keywords = [
        "ALDO", "Human growth hormone", "growth factor", "Renin"
    ]

    df_diasorin = df[df["Test"].str.contains('|'.join(liason_keywords), case=False)]
    df = df[~df["Test"].str.contains('|'.join(virology_keywords), case=False)]
    df = df[~df["Test"].str.contains('|'.join(poct_keywords), case=False)]
    df = df[~df["Test"].str.contains('|'.join(liason_keywords), case=False)]

    return df, df_diasorin

def highlight_extra_tests(book):
    raw_data_sheet = book["Raw Data"]
    cobas_sheet = book.get_sheet_by_name(" Cobas assays")

    # Create a dictionary to count occurrences in the "Raw Data" sheet
    raw_data_count = {}
    for row in raw_data_sheet.iter_rows(min_row=2, values_only=True):
        test_name = row[2] # Assuming "Test" is in the second column
        raw_data_count[test_name] = raw_data_count.get(test_name, 0) + 1

    # Create a dictionary to count occurrences in the "Cobas assays" sheet
    cobas_count = {}
    for row in cobas_sheet.iter_rows(min_row=7, values_only=True):
        test_name = row[1] # Assuming "Test" is in the second column
        cobas_count[test_name] = cobas_count.get(test_name, 0) + 1

    # Compare counts and highlight in "Raw Data" sheet if more occurrences
    for row in raw_data_sheet.iter_rows(min_row=2):
        test_cell = row[2] # Assuming "Test" is in the second column
        test_name = test_cell.value
        if raw_data_count.get(test_name, 0) > cobas_count.get(test_name, 0):
            test_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


# Get the previous month and year
today = datetime.today()
first = today.replace(day=1)
last_month = first - timedelta(days=1)
SAVE_PATH = f"{last_month.strftime('%B')}_{last_month.year}.xlsx"

# Load the CSV file and clean up the data
df = pd.read_csv(CSV_FILE_PATH, delimiter=";")
df, df_diasorin = clean_data(df)

# Load the TAE template
book = load_workbook(TAE_TEMPLATE_PATH)
raw_data_sheet = book["Raw Data"]

# Write the headers
for col_num, header in enumerate(df.columns, 1):
    raw_data_sheet.cell(row=1, column=col_num, value=header)

# Write the data
for index, row in enumerate(df.itertuples(index=False), start=2):
    for col_num, value in enumerate(row, 1):
        col_letter = get_column_letter(col_num)
        raw_data_sheet[f"{col_letter}{index}"] = value

# Create and fill Diasorin sheet
if "Diasorin" not in book.sheetnames:
    book.create_sheet("Diasorin")

diasorin_sheet = book["Diasorin"]

for col_num, header in enumerate(df_diasorin.columns, 1):
    diasorin_sheet.cell(row=1, column=col_num, value=header)

for index, row in enumerate(df_diasorin.itertuples(index=False), start=2):
    for col_num, value in enumerate(row, 1):
        col_letter = get_column_letter(col_num)
        diasorin_sheet[f"{col_letter}{index}"] = value

# Make sure at least one sheet is visible
visible_sheets = [sheet for sheet in book.worksheets if not sheet.sheet_state == 'hidden']
if not visible_sheets:
    book.active.sheet_state = 'visible'

# Highlight the extra tests
highlight_extra_tests(book)

# Save the changes to the TAE template
book.save(SAVE_PATH)
