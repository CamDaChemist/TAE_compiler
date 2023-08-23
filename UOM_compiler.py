import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import itertools

# File paths
CSV_FILE_PATH = "citm_extract.csv"
TAE_TEMPLATE_PATH = "C:/Users/Cameron.Francis/PycharmProjects/pythonProject/tae_template.xlsx"

# Get the previous month and year
today = datetime.today()
first = today.replace(day=1)
last_month = first - timedelta(days=1)
SAVE_PATH = f"{last_month.strftime('%B')}_{last_month.year}.xlsx"

# Helper function to clean up data
def clean_data(df):
    # Convert the CSV file to an ordered DataFrame
    df = pd.read_csv(CSV_FILE_PATH, delimiter=";", quotechar='"')
    df["Number of values"] = pd.to_numeric(df["Number of values"], errors="coerce")

    # Keep only the data with "Number of values" greater than 30 for tests with the same "Instrument" and "Test" but different "QC lot no."
    df = df[df["Number of values"] > 30]

    # Sort the DataFrame based on the original index
    df = df.sort_index()
    # Define Virology and POCT and Liason test keywords
    virology_keywords = [
        "CMV", "cytomegalovirus", "Epstein Barr Virus", "Hep", "hepatitis", "Herpes", "HIV",
        "Rubella", "Syphyllis", "toxoplasma", "varicella", "SYPHILIS", "Measles", "Anti-SARS-CoV-2 S Qaun"
    ]
    poct_keywords = [
        "POC", "CREATP", "INDEX"
    ]

    df = df[~df["Test"].str.contains('|'.join(virology_keywords), case=False)]
    df = df[~df["Test"].str.contains('|'.join(poct_keywords), case=False)]

    return df

# Load the CSV file and clean up the data
df = pd.read_csv(CSV_FILE_PATH, delimiter=";", quotechar='"')
df = clean_data(df)

# Load the TAE template
book = load_workbook(TAE_TEMPLATE_PATH)

# Add the DataFrame to the "Raw Data" sheet
raw_data_sheet = book["Raw Data"]

# Write the headers
for col_num, header in enumerate(df.columns, 1):
    raw_data_sheet.cell(row=1, column=col_num, value=header)

# Write the data
for index, row in enumerate(df.itertuples(index=False), start=2):
    for col_num, value in enumerate(row, 1):
        col_letter = get_column_letter(col_num)
        raw_data_sheet[f"{col_letter}{index}"] = value

# Make sure at least one sheet is visible
visible_sheets = [sheet for sheet in book.worksheets if not sheet.sheet_state == 'hidden']
if not visible_sheets:
    book.active.sheet_state = 'visible'

# Prepare to copy data to the 'UOM.' sheet
uom_sheet = book['UOM.']

data = uom_sheet.values
# Skip the next 10 rows
data = itertools.islice(data, 11, None)

uom_df = pd.DataFrame(data)

# Reset the DataFrame index
uom_df.reset_index(drop=True, inplace=True)

print(uom_df.columns)

# Prepare a mapping of column names from 'Raw Data' to 'UOM.'
column_mapping = {
    'Instrument': 'Instrument',
    'QC': 'QC',
    'QC lot No.': 'QC lot #',
    'Target mean': 'Target value',
    'SDI': 'Obtained SD',
    'Number of values': 'No. of data points (n)',
    'Calculated mean': 'Obtained Mean',
    'Calculated SD': 'Obtained SD'}
# Iterate over the Tests in the 'Raw Data' DataFrame
for test in df['Test'].unique():
    raw_test_df = df[df['Test'] == test]

    # Find the corresponding rows in the 'UOM.' DataFrame
    uom_test_df = uom_df[uom_df['Test'] == test]  # Assume 'Test' column is labeled 'Test' in 'UOM.'

    # Check if 'Raw Data' has more rows for this test than 'UOM.'
    if len(raw_test_df) > len(uom_test_df):
        # Add new rows to 'UOM.' DataFrame
        for _ in range(len(raw_test_df) - len(uom_test_df)):
            uom_df = uom_df.append(uom_test_df.iloc[0], ignore_index=True)

    # Map the 'Raw Data' to 'UOM.' DataFrame
    for raw_column, uom_column in column_mapping.items():
        uom_df.loc[uom_df['Test'] == test, uom_column] = raw_test_df[raw_column].values



# Clear the 'UOM.' sheet
for row in uom_sheet['A1:Z1000']:  # Assumes 'UOM.' sheet has 1000 rows and 26 columns
    for cell in row:
        cell.value = None

# Write the updated DataFrame back to the 'UOM.' sheet
for i, row in enumerate(uom_df.itertuples(index=False), start=1):
    for j, value in enumerate(row, start=1):
        uom_sheet.cell(row=i, column=j, value=value)

# Save the changes to the TAE template
book.save(SAVE_PATH)
